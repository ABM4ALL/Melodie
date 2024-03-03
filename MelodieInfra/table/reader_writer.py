import csv
from typing import Any, Dict, Generator, List, Tuple, Type, Union

import openpyxl
from sqlalchemy import Column, Integer, inspect
from sqlalchemy.orm import declarative_base
from sqlalchemy.types import TypeEngine

TableValues = Tuple[List[str], Generator[List[Union[str, int, float]], None, None]]
Base = declarative_base()

model_classes: Dict[str, Type[Base]] = {}


def get_stat_cls(table_name, columns: Dict[str, Column]):
    if table_name not in model_classes:
        d = {
            "_id": Column(Integer(), primary_key=True, autoincrement=True),
            "__tablename__": "{}".format(table_name),
        }
        d.update(columns)
        cls = type("MODEL_{}".format(table_name), (Base,), d)
        model_classes[table_name] = cls
    return model_classes[table_name]


class TableReader:
    def __init__(self, file_name: str = "", header=0, text_encoding="utf-8") -> None:
        self.file_name = file_name
        self.type = ""
        self.text_encoding = text_encoding
        self.header = header
        if file_name.endswith(".csv"):
            self.type = "csv"
        elif file_name.endswith((".xls", ".xlsx")):
            self.type = "excel"
        else:
            raise NotImplementedError
        self.read_methods = {"csv": self._read_csv, "excel": self._read_excel}

    def read(self):
        return self.read_methods[self.type]()

    def _read_csv(self) -> TableValues:
        f = open(self.file_name, encoding=self.text_encoding)
        reader = csv.reader(f)
        current_row = 0
        header: List[str] = []
        while current_row <= self.header:
            header = next(reader)
            current_row += 1

        def row_iter():
            for row_data in reader:
                yield row_data
            f.close()

        return header, row_iter()

    def _read_excel(self) -> TableValues:
        def excel_max_row(sheet):
            i = sheet.max_row
            real_max_row = 0
            while i > 0:
                row_dict = {i.value for i in sheet[i]}
                if row_dict == {None}:
                    i = i - 1
                else:
                    real_max_row = i
                    break

            return real_max_row

        def excel_max_col(sheet):
            i = sheet.max_column
            real_max_col = 0
            while i > 0:
                col_dict = {table.cell(row + 1, i).value for row in range(rows)}
                if col_dict == {None}:
                    i = i - 1
                else:
                    real_max_col = i
                    break

            return real_max_col

        workbook = openpyxl.load_workbook(self.file_name)
        table = workbook.active
        rows = excel_max_row(table)
        cols = excel_max_col(table)

        header = [table.cell(self.header + 1, col + 1).value for col in range(cols)]

        def row_iter():
            for row_index in range(self.header + 1, rows):
                row = [table.cell(row_index + 1, col + 1).value for col in range(cols)]
                yield row

        return header, row_iter()


# class TableSaver


class TableWriter:
    def __init__(
        self, file_name: str = "", header=0, text_encoding="utf-8", append=False
    ) -> None:
        self.file_name = file_name
        self.type = ""
        self.text_encoding = text_encoding
        self.header = header
        self.append = append
        if file_name.endswith(".csv"):
            self.type = "csv"
        elif file_name.endswith((".xls", ".xlsx")):
            self.type = "excel"
        else:
            raise NotImplementedError
        self.write_methods = {"csv": self._write_csv, "excel": self._write_excel}

    def write(self):
        return self.write_methods[self.type]()

    def _write_csv(self):
        file = open(
            self.file_name,
            "a" if self.append else "w",
            encoding=self.text_encoding,
            newline="",
        )
        writer = csv.writer(file)
        current_row = 1

        def row_writer():
            nonlocal current_row
            while 1:
                try:
                    data = yield
                    # for item in data:
                    #     file.write(str(item))
                    #     file.write(",")
                    # file.write("\an")
                    writer.writerow(data)
                    current_row += 1
                except GeneratorExit:
                    file.close()
                    return

        w = row_writer()
        next(w)
        return w

    def _write_excel(self):
        wb = openpyxl.Workbook()
        # 获取当前活跃的worksheet,默认就是第一个worksheet
        ws = wb.active
        current_row = 1

        def row_writer():
            nonlocal current_row
            while 1:
                try:
                    data = yield
                    for j, val in enumerate(data):
                        ws.cell(row=current_row, column=j + 1).value = val
                    current_row += 1
                except GeneratorExit:
                    wb.save(filename=self.file_name)
                    return

        # 保存表格
        w = row_writer()
        next(w)
        return w


class DatabaseConnector:
    def __init__(self, engine) -> None:
        self.engine = engine

    def write_table(self, table_name: str, columns, data: List[Dict]):
        stat_cls = get_stat_cls(table_name, columns)
        insp = inspect(self.engine)
        if not insp.has_table(table_name):
            stat_cls.__table__.create(bind=self.engine)
        sql = stat_cls.__table__.insert()
        with self.engine.connect() as conn:
            conn.execute(sql, data)

    def read_sql(
        self, table_name: str, sql: str
    ) -> Tuple[List[Dict[str, Any]], Dict[str, TypeEngine]]:
        # 创建inspector对象
        insp = inspect(self.engine)
        columns = insp.get_columns(table_name)
        index_mapping = []
        types = {}
        for column in columns:
            column_name = column["name"]
            column_type = column["type"]
            types[column_name] = column_type
            index_mapping.append(column_name)

        with self.engine.connect() as conn:
            result = conn.execute(sql)
            data = []
            for row in result.fetchall():
                data.append({index_mapping[i]: item for i, item in enumerate(row)})
            return data, types
